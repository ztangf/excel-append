import os
import pandas as pd
from openpyxl import load_workbook
import locale

#locale.setlocale(locale.LC_ALL, 'pt_BR')

current_dir = os.path.dirname(__file__)

files = os.listdir(current_dir)

#file_path = 'C:\Users\Rodrigo Dácio\Documents\MyCode\Projects\Test\data_test_excel_mod.xlsx'


#wb = load_workbook('data_test_excel_mod.xlsx')
#ws = wb['Sheet1']

# wb = load_workbook('new_Consumo_mod2.xlsx')
# ws = wb['Consumo_água']


#Lendo planilha
df = pd.read_excel('Consumo água_01.xlsx')

#Não Funciona
#df['MÊS DE PAGAMENTO'] = pd.to_datetime(df['MÊS DE PAGAMENTO'], format = '%m/%Y' )
    
#last_row = ws.max_row

#data = {"Consumo": [4537] ,"Dias": [25]}
data = {"MÊS DE PAGAMENTO" :['12/01/2024']}
new_df = pd.DataFrame(data)
df = pd.concat([df,new_df],ignore_index=True)

data2 = {"Consumo": [8000] ,"Dias": [21]}
new_df2 = pd.DataFrame(data2)
df = pd.concat([df,new_df2],ignore_index=True)

df['Teste Formula'] = df['Consumo'] + df['Dias']


#Estilo Colunas
#Mes correto 
df['MÊS DE PAGAMENTO'] = pd.to_datetime(df['MÊS DE PAGAMENTO'],dayfirst=True)
df['MÊS DE PAGAMENTO'] = df['MÊS DE PAGAMENTO'].dt.strftime('%b/%Y')

#Moeda


# #Dessa forma consigo imputar os dados selecionando a coluna e mantendo o padrão
# for r, row in enumerate(df.values.tolist(), 2):
#     for c, value in enumerate(row,start=1):
#         ws.cell(row=r,column=c).value = value

# wb.save('new_Consumo_mod3.xlsx')

#Desta forma altera o padrão do excel
df.to_excel('new_Consumo_mod2.xlsx', index = False)
